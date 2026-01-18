"""
Admin API endpoints - PostgreSQL con SQLAlchemy
"""
import json
from fastapi import APIRouter, HTTPException
from pydantic import BaseModel
from typing import Optional, List
from database import get_config, set_config, get_all_config, is_tool_enabled, set_tool_enabled, get_all_tools_config
from models import SessionLocal, Cita, Inventario, Memoria, Disponibilidad, HorarioBloqueado

router = APIRouter(prefix="/api", tags=["admin"])

DIAS_SEMANA = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]


# Pydantic Models
class ApiKeysModel(BaseModel):
    openai_api_key: Optional[str] = ""
    whatsapp_api_url: Optional[str] = ""
    whatsapp_api_key: Optional[str] = ""
    whatsapp_session: Optional[str] = ""


class PromptModel(BaseModel):
    system_prompt: Optional[str] = ""
    prompt_sections: Optional[dict] = None
    model: Optional[str] = "gpt-4o-mini"
    temperature: Optional[float] = 0.7
    max_tokens: Optional[int] = 500
    business_name: Optional[str] = ""
    business_type: Optional[str] = ""


class ImprovePromptModel(BaseModel):
    section: str
    current_content: Optional[str] = ""
    business_name: Optional[str] = ""
    business_type: Optional[str] = ""
    all_sections: Optional[dict] = None


class ToolToggle(BaseModel):
    enabled: bool


class ProductModel(BaseModel):
    producto: str
    stock: int
    precio: float




# Stats
@router.get("/stats")
async def get_stats():
    db = SessionLocal()
    try:
        conversations = db.query(Memoria).count()
        appointments = db.query(Cita).count()
        products = db.query(Inventario).count()
        
        memorias = db.query(Memoria).all()
        total_messages = 0
        for m in memorias:
            if m.historial:
                try:
                    historial = json.loads(m.historial)
                    total_messages += len(historial)
                except:
                    pass
        
        return {
            "totalConversations": conversations,
            "totalAppointments": appointments,
            "totalProducts": products,
            "totalMessages": total_messages,
        }
    finally:
        db.close()


# API Keys
@router.get("/config/api-keys")
async def get_api_keys():
    openai_key = get_config("openai_api_key", "")
    whatsapp_key = get_config("whatsapp_api_key", "")
    return {
        "openai_api_key": openai_key[:8] + "..." if openai_key else "",
        "whatsapp_api_url": get_config("whatsapp_api_url", ""),
        "whatsapp_api_key": whatsapp_key[:8] + "..." if whatsapp_key else "",
        "whatsapp_session": get_config("whatsapp_session", ""),
    }


@router.put("/config/api-keys")
async def update_api_keys(keys: ApiKeysModel):
    data = keys.dict(exclude_unset=True)
    for key, value in data.items():
        if value and not value.endswith("..."):
            set_config(key, value)
    return {"status": "ok"}


# Prompt
@router.get("/prompt")
async def get_prompt():
    prompt_sections_str = get_config("prompt_sections", "")
    prompt_sections = None
    if prompt_sections_str:
        try:
            prompt_sections = json.loads(prompt_sections_str)
        except:
            pass
    
    return {
        "system_prompt": get_config("system_prompt", ""),
        "prompt_sections": prompt_sections,
        "model": get_config("model", "gpt-4o-mini"),
        "temperature": float(get_config("temperature", "0.7")),
        "max_tokens": int(get_config("max_tokens", "500")),
        "business_name": get_config("business_name", ""),
        "business_type": get_config("business_type", ""),
    }


@router.put("/prompt")
async def update_prompt(prompt: PromptModel):
    data = prompt.dict()
    for key, value in data.items():
        if value is not None:
            if key == "prompt_sections":
                set_config(key, json.dumps(value) if value else "")
            else:
                set_config(key, str(value))
    return {"status": "ok"}


@router.post("/prompt/improve")
async def improve_prompt(data: ImprovePromptModel):
    """Usa IA para mejorar una sección del prompt"""
    import os
    from openai import OpenAI
    
    api_key = get_config("openai_api_key", "") or os.getenv("OPENAI_API_KEY", "")
    if not api_key:
        return {"error": "No hay API key de OpenAI configurada"}
    
    client = OpenAI(api_key=api_key)
    
    section_names = {
        "role": "Rol (quién es el agente)",
        "context": "Contexto (situación del negocio)", 
        "task": "Tarea (objetivos principales)",
        "constraints": "Restricciones (reglas y límites)",
        "tone": "Tono (cómo comunicarse)"
    }
    
    if data.section == "all":
        # Mejorar todas las secciones
        improved_sections = {}
        for section_key in ["role", "context", "task", "constraints", "tone"]:
            current = data.all_sections.get(section_key, "") if data.all_sections else ""
            
            prompt = f"""Mejora el siguiente texto para un prompt de IA de atención al cliente.
Negocio: {data.business_name} ({data.business_type})
Sección: {section_names.get(section_key, section_key)}

Texto actual:
{current}

Instrucciones:
- Mantén la esencia pero hazlo más efectivo
- Sé específico y claro
- Usa un estilo profesional
- Máximo 3-4 oraciones
- Responde SOLO con el texto mejorado, sin explicaciones"""

            response = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[{"role": "user", "content": prompt}],
                temperature=0.7,
                max_tokens=300
            )
            improved_sections[section_key] = response.choices[0].message.content.strip()
        
        return {"improved_sections": improved_sections}
    else:
        # Mejorar una sección específica
        section_name = section_names.get(data.section, data.section)
        
        # Contexto de otras secciones
        context_parts = []
        if data.all_sections:
            for key, name in section_names.items():
                if key != data.section and data.all_sections.get(key):
                    context_parts.append(f"{name}: {data.all_sections[key][:100]}...")
        
        context_str = "\n".join(context_parts) if context_parts else "No hay otras secciones definidas"
        
        prompt = f"""Mejora el siguiente texto para un prompt de IA de atención al cliente.

Negocio: {data.business_name} ({data.business_type})
Sección a mejorar: {section_name}

Contexto del resto del prompt:
{context_str}

Texto actual a mejorar:
{data.current_content}

Instrucciones:
- Mantén la esencia pero hazlo más efectivo y profesional
- Sé específico y detallado
- Considera el contexto del negocio
- Máximo 4-5 oraciones
- Responde SOLO con el texto mejorado, sin explicaciones ni comillas"""

        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.7,
            max_tokens=300
        )
        
        return {"improved": response.choices[0].message.content.strip()}


# Tools
@router.get("/tools")
async def get_tools():
    tools = get_all_tools_config()
    return [
        {"id": t["nombre"], "enabled": t["habilitado"], "description": t["descripcion"]}
        for t in tools
    ]


@router.patch("/tools/{name}")
async def toggle_tool(name: str, data: ToolToggle):
    set_tool_enabled(name, data.enabled)
    return {"status": "ok"}


# Inventory
@router.get("/inventory")
async def get_inventory():
    db = SessionLocal()
    try:
        products = db.query(Inventario).all()
        return [p.to_dict() for p in products]
    finally:
        db.close()


@router.post("/inventory")
async def create_product(product: ProductModel):
    db = SessionLocal()
    try:
        new_product = Inventario(
            producto=product.producto,
            stock=product.stock,
            precio=product.precio
        )
        db.add(new_product)
        db.commit()
        db.refresh(new_product)
        return new_product.to_dict()
    finally:
        db.close()


@router.put("/inventory/{product_id}")
async def update_product(product_id: int, product: ProductModel):
    db = SessionLocal()
    try:
        p = db.query(Inventario).filter(Inventario.id == product_id).first()
        if p:
            p.producto = product.producto
            p.stock = product.stock
            p.precio = product.precio
            db.commit()
        return {"status": "ok"}
    finally:
        db.close()


@router.delete("/inventory/{product_id}")
async def delete_product(product_id: int):
    db = SessionLocal()
    try:
        db.query(Inventario).filter(Inventario.id == product_id).delete()
        db.commit()
        return {"status": "ok"}
    finally:
        db.close()


from fastapi import UploadFile, File

@router.post("/inventory/upload")
async def upload_inventory(file: UploadFile = File(...)):
    """Procesa un archivo CSV o Excel"""
    import os
    import csv
    import io
    import re
    from openai import OpenAI
    
    content = await file.read()
    filename = file.filename.lower() if file.filename else ""
    
    rows = []
    
    # Procesar Excel
    if filename.endswith('.xlsx') or filename.endswith('.xls'):
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(filename=io.BytesIO(content), read_only=True)
            sheet = workbook.active
            for row in sheet.iter_rows(values_only=True):
                # Convertir None a string vacío y todo a string
                rows.append([str(cell) if cell is not None else '' for cell in row])
            workbook.close()
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error al leer Excel: {str(e)}")
    
    # Procesar CSV
    else:
        try:
            text_content = content.decode('utf-8')
        except:
            text_content = content.decode('latin-1')
        
        lines = [l.strip() for l in text_content.strip().split('\n') if l.strip()]
        if not lines:
            raise HTTPException(status_code=400, detail="El archivo está vacío")
        
        # Detectar delimitador
        delimiter = ','
        if ';' in lines[0] and ',' not in lines[0]:
            delimiter = ';'
        elif '\t' in lines[0]:
            delimiter = '\t'
        
        reader = csv.reader(io.StringIO(text_content), delimiter=delimiter)
        rows = list(reader)
    
    # Validar que hay datos
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="El archivo necesita al menos una cabecera y una fila de datos")
    
    headers = [h.strip().lower() for h in rows[0]]
    data_rows = rows[1:]
    
    print(f"[Upload] Cabeceras detectadas: {headers}")
    print(f"[Upload] Filas de datos: {len(data_rows)}")
    if data_rows:
        print(f"[Upload] Primera fila: {data_rows[0]}")
    
    # Usar IA solo para mapear cabeceras desconocidas
    mapping = {}
    api_key = get_config("openai_api_key", "") or os.getenv("OPENAI_API_KEY", "")
    if api_key:
        client = OpenAI(api_key=api_key)
        mapping_prompt = f"""Analiza estas cabeceras de un archivo de inventario y mapéalas a los campos estándar.

Cabeceras encontradas: {headers}

Responde SOLO con un JSON object que mapee cada cabecera al campo correspondiente.
Campos válidos: producto, stock, precio, categoria, descripcion
Si una cabecera no corresponde a ningún campo, usa null.

Ejemplo: {{"nombre": "producto", "cantidad": "stock", "costo": "precio", "tipo": "categoria", "notas": "descripcion", "codigo": null}}

RESPONDE SOLO EL JSON, SIN TEXTO ADICIONAL."""

        try:
            response = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[{"role": "user", "content": mapping_prompt}],
                temperature=0,
                max_tokens=200
            )
            result = response.choices[0].message.content.strip()
            if result.startswith('```'):
                result = result.split('```')[1].split('```')[0]
                if result.startswith('json'):
                    result = result[4:]
            mapping = json.loads(result.strip())
        except:
            mapping = {}
    
    # Mapeo por defecto basado en nombres comunes
    default_mapping = {
        'producto': 'producto', 'nombre': 'producto', 'product': 'producto', 'name': 'producto', 
        'item': 'producto', 'articulo': 'producto', 'servicio': 'producto', 'descripcion_producto': 'producto',
        'stock': 'stock', 'cantidad': 'stock', 'qty': 'stock', 'quantity': 'stock', 'existencia': 'stock', 'inventario': 'stock',
        'precio': 'precio', 'price': 'precio', 'costo': 'precio', 'cost': 'precio', 'valor': 'precio', 'importe': 'precio',
        'categoria': 'categoria', 'category': 'categoria', 'tipo': 'categoria', 'type': 'categoria', 'grupo': 'categoria',
        'descripcion': 'descripcion', 'description': 'descripcion', 'detalle': 'descripcion', 'notas': 'descripcion'
    }
    
    # Combinar mapeos (IA tiene prioridad)
    final_mapping = {}
    for h in headers:
        if h in mapping and mapping[h]:
            final_mapping[h] = mapping[h]
        elif h in default_mapping:
            final_mapping[h] = default_mapping[h]
    
    print(f"[Upload] Mapeo IA: {mapping}")
    print(f"[Upload] Mapeo final: {final_mapping}")
    
    # Extraer productos programáticamente
    products = []
    for row in data_rows:
        if len(row) < len(headers):
            row.extend([''] * (len(headers) - len(row)))
        
        product = {'producto': '', 'stock': 10, 'precio': 0, 'categoria': '', 'descripcion': ''}
        
        for i, header in enumerate(headers):
            if header in final_mapping:
                field = final_mapping[header]
                value = row[i].strip() if i < len(row) else ''
                
                if field == 'stock':
                    try:
                        product['stock'] = int(float(re.sub(r'[^\d.]', '', value) or 0))
                    except:
                        product['stock'] = 10
                elif field == 'precio':
                    try:
                        product['precio'] = float(re.sub(r'[^\d.]', '', value) or 0)
                    except:
                        product['precio'] = 0
                else:
                    product[field] = value
        
        if product['producto']:
            products.append(product)
    
    if not products:
        raise HTTPException(status_code=400, detail="No se encontraron productos. Verifica que haya una columna con nombres de productos.")
    
    return {"preview": products, "mapping": final_mapping, "headers": headers}


class ImportProductsModel(BaseModel):
    products: List[dict]

@router.post("/inventory/import")
async def import_inventory(data: ImportProductsModel):
    """Importa productos en masa"""
    db = SessionLocal()
    try:
        imported = 0
        for item in data.products:
            new_product = Inventario(
                producto=item.get('producto', ''),
                stock=int(item.get('stock', 0)),
                precio=float(item.get('precio', 0))
            )
            db.add(new_product)
            imported += 1
        
        db.commit()
        return {"success": True, "imported": imported}
    except Exception as e:
        db.rollback()
        return {"success": False, "error": str(e)}
    finally:
        db.close()


# Appointments
@router.get("/appointments")
async def get_appointments():
    db = SessionLocal()
    try:
        citas = db.query(Cita).order_by(Cita.fecha.desc(), Cita.hora.desc()).all()
        return [c.to_dict() for c in citas]
    finally:
        db.close()


@router.patch("/appointments/{appointment_id}/status")
async def update_appointment_status(appointment_id: int, data: dict):
    db = SessionLocal()
    try:
        cita = db.query(Cita).filter(Cita.id == appointment_id).first()
        if cita:
            cita.estado = data.get("estado", "pendiente")
            db.commit()
        return {"status": "ok"}
    finally:
        db.close()


@router.delete("/appointments/{appointment_id}")
async def delete_appointment(appointment_id: int):
    db = SessionLocal()
    try:
        db.query(Cita).filter(Cita.id == appointment_id).delete()
        db.commit()
        return {"status": "ok"}
    finally:
        db.close()


# Availability
@router.get("/availability")
async def get_availability():
    db = SessionLocal()
    try:
        disponibilidad = db.query(Disponibilidad).order_by(Disponibilidad.dia_semana).all()
        return [d.to_dict() for d in disponibilidad]
    finally:
        db.close()


@router.put("/availability/{dia_id}")
async def update_availability(dia_id: int, data: dict):
    db = SessionLocal()
    try:
        disp = db.query(Disponibilidad).filter(Disponibilidad.id == dia_id).first()
        if disp:
            disp.hora_inicio = data.get("hora_inicio", disp.hora_inicio)
            disp.hora_fin = data.get("hora_fin", disp.hora_fin)
            disp.activo = data.get("activo", disp.activo)
            db.commit()
        return {"status": "ok"}
    finally:
        db.close()


# Blocked Slots
@router.get("/blocked-slots")
async def get_blocked_slots():
    db = SessionLocal()
    try:
        slots = db.query(HorarioBloqueado).order_by(HorarioBloqueado.fecha, HorarioBloqueado.hora).all()
        return [s.to_dict() for s in slots]
    finally:
        db.close()


@router.post("/blocked-slots")
async def add_blocked_slot(data: dict):
    db = SessionLocal()
    try:
        slot = HorarioBloqueado(
            fecha=data.get("fecha"),
            hora=data.get("hora"),
            motivo=data.get("motivo", "Bloqueado")
        )
        db.add(slot)
        db.commit()
        db.refresh(slot)
        return {"status": "ok", "id": slot.id}
    finally:
        db.close()


@router.delete("/blocked-slots/{slot_id}")
async def delete_blocked_slot(slot_id: int):
    db = SessionLocal()
    try:
        db.query(HorarioBloqueado).filter(HorarioBloqueado.id == slot_id).delete()
        db.commit()
        return {"status": "ok"}
    finally:
        db.close()


# Available slots for date
@router.get("/available-slots/{fecha}")
async def get_available_slots(fecha: str):
    from tools import obtener_horarios_disponibles
    return {"fecha": fecha, "horarios": obtener_horarios_disponibles(fecha)}


# Conversations
@router.get("/conversations")
async def get_conversations():
    db = SessionLocal()
    try:
        memorias = db.query(Memoria).order_by(Memoria.updated_at.desc()).all()
        result = []
        for m in memorias:
            historial = json.loads(m.historial) if m.historial else []
            ultimo = historial[-1]["content"] if historial else ""
            result.append({
                "telefono": m.telefono,
                "ultimo_mensaje": ultimo[:50] + "..." if len(ultimo) > 50 else ultimo,
                "fecha": m.updated_at.isoformat() if m.updated_at else "",
                "mensajes_count": len(historial),
            })
        return result
    finally:
        db.close()


@router.get("/conversations/{phone}")
async def get_conversation(phone: str):
    db = SessionLocal()
    try:
        memoria = db.query(Memoria).filter(Memoria.telefono == phone).first()
        if memoria and memoria.historial:
            return {"messages": json.loads(memoria.historial)}
        return {"messages": []}
    finally:
        db.close()


@router.delete("/conversations/{phone}")
async def delete_conversation(phone: str):
    db = SessionLocal()
    try:
        db.query(Memoria).filter(Memoria.telefono == phone).delete()
        db.commit()
        return {"status": "ok"}
    finally:
        db.close()


