"""
Inventory router
"""
import os
import csv
import io
import re
import json
from typing import List
from fastapi import APIRouter, HTTPException, UploadFile, File, Depends
from openai import OpenAI

from models import SessionLocal, Inventario, Usuario
from database import get_config
from api.schemas.inventory import ProductModel, ImportProductsModel
from auth import get_current_user

router = APIRouter(prefix="/inventory", tags=["inventory"])


@router.get("/")
async def get_inventory(current_user: Usuario = Depends(get_current_user)):
    """Get all products"""
    db = SessionLocal()
    try:
        products = db.query(Inventario).all()
        return [p.to_dict() for p in products]
    finally:
        db.close()


@router.post("/")
async def create_product(product: ProductModel, current_user: Usuario = Depends(get_current_user)):
    """Create new product"""
    db = SessionLocal()
    try:
        new_product = Inventario(
            producto=product.producto,
            stock=product.stock,
            precio=product.precio
        )
        db.add(new_product)
        db.commit()
        db.refresh(new_product)
        return new_product.to_dict()
    finally:
        db.close()


@router.put("/{product_id}")
async def update_product(product_id: int, product: ProductModel, current_user: Usuario = Depends(get_current_user)):
    """Update product"""
    db = SessionLocal()
    try:
        p = db.query(Inventario).filter(Inventario.id == product_id).first()
        if p:
            p.producto = product.producto
            p.stock = product.stock
            p.precio = product.precio
            db.commit()
        return {"status": "ok"}
    finally:
        db.close()


@router.delete("/{product_id}")
async def delete_product(product_id: int, current_user: Usuario = Depends(get_current_user)):
    """Delete product"""
    db = SessionLocal()
    try:
        db.query(Inventario).filter(Inventario.id == product_id).delete()
        db.commit()
        return {"status": "ok"}
    finally:
        db.close()


@router.post("/upload")
async def upload_inventory(file: UploadFile = File(...), current_user: Usuario = Depends(get_current_user)):
    """Process CSV or Excel file"""
    content = await file.read()
    filename = file.filename.lower() if file.filename else ""
    
    rows = []
    
    # Process Excel
    if filename.endswith('.xlsx') or filename.endswith('.xls'):
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(filename=io.BytesIO(content), read_only=True)
            sheet = workbook.active
            for row in sheet.iter_rows(values_only=True):
                rows.append([str(cell) if cell is not None else '' for cell in row])
            workbook.close()
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error reading Excel: {str(e)}")
    
    # Process CSV
    else:
        try:
            text_content = content.decode('utf-8')
        except:
            text_content = content.decode('latin-1')
        
        lines = [l.strip() for l in text_content.strip().split('\n') if l.strip()]
        if not lines:
            raise HTTPException(status_code=400, detail="File is empty")
        
        delimiter = ','
        if ';' in lines[0] and ',' not in lines[0]:
            delimiter = ';'
        elif '\t' in lines[0]:
            delimiter = '\t'
        
        reader = csv.reader(io.StringIO(text_content), delimiter=delimiter)
        rows = list(reader)
    
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="File needs at least header and one data row")
    
    headers = [h.strip().lower() for h in rows[0]]
    data_rows = rows[1:]
    
    # Use AI only to map unknown headers
    mapping = {}
    api_key = get_config("openai_api_key", "") or os.getenv("OPENAI_API_KEY", "")
    if api_key:
        try:
            client = OpenAI(api_key=api_key)
            mapping_prompt = f"""Map these inventory file headers to standard fields.

Headers found: {headers}

Respond ONLY with a JSON object mapping each header to the corresponding field.
Valid fields: producto, stock, precio, categoria, descripcion
If a header doesn't match any field, use null.

Example: {{"name": "producto", "quantity": "stock", "price": "precio", "type": "categoria", "notes": "descripcion", "code": null}}

RESPOND ONLY THE JSON, NO ADDITIONAL TEXT."""

            response = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[{"role": "user", "content": mapping_prompt}],
                temperature=0,
                max_tokens=200
            )
            result = response.choices[0].message.content.strip()
            if result.startswith('```'):
                result = result.split('```')[1].split('```')[0]
                if result.startswith('json'):
                    result = result[4:]
            mapping = json.loads(result.strip())
        except:
            mapping = {}
    
    # Default mapping for common names
    default_mapping = {
        'producto': 'producto', 'nombre': 'producto', 'product': 'producto', 'name': 'producto', 
        'item': 'producto', 'articulo': 'producto', 'servicio': 'producto',
        'stock': 'stock', 'cantidad': 'stock', 'qty': 'stock', 'quantity': 'stock',
        'precio': 'precio', 'price': 'precio', 'costo': 'precio', 'cost': 'precio',
        'categoria': 'categoria', 'category': 'categoria', 'tipo': 'categoria',
        'descripcion': 'descripcion', 'description': 'descripcion', 'detalle': 'descripcion'
    }
    
    final_mapping = {}
    for h in headers:
        if h in mapping and mapping[h]:
            final_mapping[h] = mapping[h]
        elif h in default_mapping:
            final_mapping[h] = default_mapping[h]
    
    # Extract products programmatically
    products = []
    for row in data_rows:
        if len(row) < len(headers):
            row.extend([''] * (len(headers) - len(row)))
        
        product = {'producto': '', 'stock': 10, 'precio': 0, 'categoria': '', 'descripcion': ''}
        
        for i, header in enumerate(headers):
            if header in final_mapping:
                field = final_mapping[header]
                value = row[i].strip() if i < len(row) else ''
                
                if field == 'stock':
                    try:
                        product['stock'] = int(float(re.sub(r'[^\d.]', '', value) or 0))
                    except:
                        product['stock'] = 10
                elif field == 'precio':
                    try:
                        product['precio'] = float(re.sub(r'[^\d.]', '', value) or 0)
                    except:
                        product['precio'] = 0
                else:
                    product[field] = value
        
        if product['producto']:
            products.append(product)
    
    if not products:
        raise HTTPException(status_code=400, detail="No products found. Check that there's a product name column.")
    
    return {"preview": products, "mapping": final_mapping, "headers": headers}


@router.post("/import")
async def import_inventory(data: ImportProductsModel, current_user: Usuario = Depends(get_current_user)):
    """Import products in bulk"""
    db = SessionLocal()
    try:
        imported = 0
        for item in data.products:
            new_product = Inventario(
                producto=item.get('producto', ''),
                stock=int(item.get('stock', 0)),
                precio=float(item.get('precio', 0))
            )
            db.add(new_product)
            imported += 1
        
        db.commit()
        return {"success": True, "imported": imported}
    except Exception as e:
        db.rollback()
        return {"success": False, "error": str(e)}
    finally:
        db.close()
